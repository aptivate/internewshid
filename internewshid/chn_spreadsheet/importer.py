import datetime
from decimal import Decimal

from django.utils import six
from django.utils.timezone import is_naive
from django.utils.translation import ugettext as _

import dateutil.parser
import pytz
from openpyxl import load_workbook

import transport
from transport.exceptions import ItemNotUniqueException, TransportException

from .models import SheetProfile


class SheetImportException(Exception):
    pass


class Importer(object):

    def __init__(self):
        self.profile = None

    def get_profile(self, label):
        try:
            sheet_profile = SheetProfile.objects.get(label=label)
        except SheetProfile.DoesNotExist:
            error_msg = _('Misconfigured service. Source "{0}" does not exist').format(label)
            raise SheetImportException(error_msg)

        return sheet_profile.profile

    def get_columns_map(self):
        '''This function assumes that column names are unique for spreadsheet.
        If they are not, then you already have a problem.'''

        return {column['name']: column for column in self.profile['columns']}

    def get_rows_iterator(self, spreadsheet, file_format):
        if file_format == 'excel':
            try:
                wb = load_workbook(spreadsheet, read_only=True)
                ws = wb[wb.sheetnames[0]]
            except Exception:
                error_msg = _('Expected excel file. Received file in an unrecognized format.')
                raise SheetImportException(error_msg)
            rows = ws.rows
        else:
            error_msg = _('Unsupported file format: {0}').format(file_format)
            raise SheetImportException(error_msg)

        return rows

    def order_columns(self, first_row=None):
        profile_columns = self.profile['columns']

        columns = []
        if first_row:
            col_map = self.get_columns_map()

            for label in first_row[:len(col_map)]:
                if label is not None:
                    try:
                        stripped_label = label.strip()
                        columns.append(col_map[stripped_label])
                    except Exception:
                        error_msg = _('Unknown column: {0}').format(label)
                        raise SheetImportException(error_msg)
        else:
            columns = [d.copy() for d in profile_columns]

        return columns

    def get_fields_and_types(self, columns):
        fields = [col['field'] for col in columns]
        types = [col['type'] for col in columns]

        return fields, types

    def normalize_row(self, raw_row):
        # Unify difference between CSV and openpyxl cells
        return [getattr(v, 'value', v) for v in raw_row]

    def process_rows(self, rows):
        skip_header = self.profile.get('skip_header', False)
        meta_data = self.profile.get('taxonomies')

        # If there is no header (skip_header=False), then use profile's order of
        # columns, otherwise use header line to check mapping and define order
        first_row = self.normalize_row(next(rows)) if skip_header else None
        columns = self.order_columns(first_row)

        objects = []
        for i, row in enumerate(rows, 2 if first_row else 1):
            try:
                values = self.normalize_row(row)

                if any(values):
                    item = self.process_row(values, columns)
                    item['_row_number'] = i
                    for taxonomy, term in meta_data.items():
                        self._append_term_to_item(item, taxonomy, term)

                    objects.append(item)

            except SheetImportException as e:
                raise type(e)(str(e) + 'in row {0} '.format(i)) from e

        return objects

    def process_row(self, values, columns):
        item = {}

        for val, col in zip(values, columns):
            converter = CellConverter(val, col)

            if col['type'] == 'taxonomy':
                value = converter.convert_value()
                for term in value.split(','):
                    self._append_term_to_item(
                        item, col['taxonomy'], term.strip())
            elif col['type'] == 'protection_concern':
                value = converter.convert_value()
                if value:
                    self._append_term_to_item(
                        item, 'tags', value.strip())
            else:
                converter.add_to(item)

        return item

    def _append_term_to_item(self, item, taxonomy, name):
        term = self._get_term_dict(taxonomy, name)
        item.setdefault('terms', []).append(term)

    def _get_term_dict(self, taxonomy, name):
        return {'taxonomy': taxonomy, 'name': name}

    def save_rows(self, objects):
        num_saved = 0

        for obj in objects:
            row = obj.pop('_row_number', '')
            terms = obj.pop('terms', [])
            try:
                item = transport.items.create(obj)
                for term in terms:
                    transport.items.add_terms(
                        item['id'], term['taxonomy'], term['name'])
            except ItemNotUniqueException:
                pass

            except TransportException as exc_inst:
                message = self._get_spreadsheet_error_message(row, exc_inst)

                raise SheetImportException(message)

            else:
                num_saved += 1

        return num_saved

    def _get_spreadsheet_error_message(self, row, exc_inst):
        # The exception will have different formats:
        # Example of failure during transport.items.create():
        # {
        #  'contributor': [ErrorDetail(string='Ensure this field has no more than 190 characters.', code='max_length')],
        #  'status_code': 400,
        #  'item': {
        #      'timestamp': datetime.datetime(2018, 8, 9, 12, 14, 26, 766000, tzinfo=tzoffset(None, 21600)),
        #      'body': 'the community members want more food.',
        #      'gender': 'female',
        #      'location': 'Camp 4',
        #      'contributor': '<a very long string>',
        #      'external_id': '97f61035-6feb-40a1-9e7e-15c0f65cfdb5'
        #   }
        # }

        # Example of failure during transport.items.add_terms():
        # {
        #  'detail': 'Term matching query does not exist.',
        #  'status_code': 400,
        #  'terms': {'taxonomy': 'age-ranges', 'name': ''},
        #  'item_id': 4
        # }

        exc_inst.message.pop('status_code', None)
        exc_inst.message.pop('item_id', None)
        item = exc_inst.message.pop('item', {})
        detail = exc_inst.message.pop('detail', None)
        terms = exc_inst.message.pop('terms', None)

        messages = []

        field_to_column_map = self._get_field_to_column_map()

        for field, errors in exc_inst.message.items():
            for error in errors:
                messages.append(
                    _("Column: '{0}' ({1})\nError ({2}): '{3}'\n\nValue: {4}").format(
                        field_to_column_map.get(field),
                        field,
                        getattr(error, 'code', ''),
                        six.text_type(error),
                        item.get(field, '')
                    )
                )

        if detail and terms:
            taxonomy = terms['taxonomy']
            name = terms['name']

            messages.append(
                _(f"Error: {detail}\nTaxonomy: {taxonomy}\nName: {name}\n")
            )

        return _("There was a problem with row {0} of the spreadsheet:\n{1}").format(
            row, '\n'.join(messages)
        )

    def _get_field_to_column_map(self):
        return {column['field']: column['name'] for column in self.profile['columns']}

    def store_spreadsheet(self, label, fobject):
        self.profile = self.get_profile(label)

        file_format = self.profile.get('format')

        rows = self.get_rows_iterator(fobject, file_format)

        items = self.process_rows(rows)

        num_saved = self.save_rows(items)

        return (num_saved, len(items) - num_saved)


class CellConverter(object):

    def __init__(self, value, col_spec):
        self.value = value
        self.type = col_spec['type']
        self.field = col_spec['field']
        self.date_format = col_spec.get('date_format', None)

    def add_to(self, object_dict):
        if self.type != 'ignore':
            object_dict[self.field] = self.convert_value()
        return object_dict

    def convert_value(self):
        converters = {
            'date': lambda x: self.convert_date(),
            'text': lambda x: x if x else '',
            'integer': lambda x: int(x),
            'number': lambda x: Decimal(x),
            'taxonomy': lambda x: x if x else '',
            'protection_concern': lambda x: 'Protection Concern' if x.lower() == 'yes' else ''
        }
        if self.type not in converters:
            raise SheetImportException(
                _(u"Unknown data type '{0}' ").format(self.type))
        try:
            return converters[self.type](self.value)
        except Exception as e:
            message = _("{0}: Can not process value '{1}' of type '{2}' ").format(str(e), self.value, self.type)
            raise SheetImportException(message) from e

    def convert_date(self):
        if self.value is None:
            return None

        if isinstance(self.value, str):
            date_time = self.parse_date()
        else:
            date_time = self.value

        if is_naive(date_time):
            date_time = pytz.utc.localize(date_time)

        return date_time

    def parse_date(self):
        if self.date_format is None:
            raise SheetImportException(
                _(u"Date format not specified for '{0}' ").format(self.field))

        try:
            date_time = datetime.datetime.strptime(self.value,
                                                   self.date_format)
        except Exception:
            date_time = dateutil.parser.parse(self.value)

        return date_time
