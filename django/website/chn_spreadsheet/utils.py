import dateutil.parser
from decimal import Decimal
import sys

from django.utils.translation import ugettext as _
from openpyxl import load_workbook

from transport.data_layer_transport import create_message

from .models import SheetProfile


class SheetImportException(Exception):
    pass


def get_profile(label):
    try:
        sheet_profile = SheetProfile.objects.get(label=label)
    except SheetProfile.DoesNotExist:
        error_msg = _('Misconfigured service. Source "%s" does not exist') % label
        raise SheetImportException(error_msg)

    # TODO: Revert to using database
    # return sheet_profile.profile

    return {
        "label": "geopoll",
        "name": "Geopoll",
        "format": "excel",
        "type": "message",
        "columns": [
            {
                "name": "Province",
                "type": "ignore",
                "field": "ignore"
            },
            {
                "name": "CreatedDate",
                "type": "date",
                "field": "timestamp"
            },
            {
                "name": "AgeGroup",
                "type": "ignore",
                "field": "ignore"
            },
            {
                "name": "QuestIO",
                "type": "text",
                "field": "body"
            }
        ],
        "skip_header": 1
    }


def get_columns_map(col_list):
    '''This function assumes that column names are unique for spreadsheet.
    If they are not, then you already have a problem.'''

    # Python 2.7 (should be faster than a loop)
    cols = {
        column['name']: {
            'type': column['type'],
            'field': column['field']
        } for column in col_list
    }
    return cols


def get_rows_iterator(spreadsheet, file_format):
    if file_format == 'excel':
        try:
            wb = load_workbook(spreadsheet, read_only=True)
            ws = wb[wb.sheetnames[0]]
        except:
            error_msg = _('Expected excel file. Received file in an unrecognized format.')
            raise SheetImportException(error_msg)
        rows = ws.rows
    else:
        error_msg = _('Unsupported file format: %s') % file_format
        raise SheetImportException(error_msg)
    return rows


def order_columns(profile_columns, first_row=None):
    columns = []
    if first_row:
        col_map = get_columns_map(profile_columns)
        for label in first_row:
            try:
                columns.append(col_map[label])
            except:
                error_msg = _('Unknown column: %s') % label
                raise SheetImportException(error_msg)
    else:
        columns = [d.copy() for d in profile_columns]
        for col in columns:
            del col['name']  # Unify with first row version

    return columns


def get_fields_and_types(columns):
    fields = [col['field'] for col in columns]
    types = [col['type'] for col in columns]
    return fields, types


def parse_date(value):
    if isinstance(value, basestring):
        date_time = dateutil.parser.parse(value, dayfirst=True)
    else:
        date_time = value

    return date_time.date()


def convert_value(value, type):
    converters = {
        'text': lambda x: x,
        'date': parse_date,
        'integer': lambda x: int(x),
        'number': lambda x: Decimal(x)
    }
    if type not in converters:
        raise SheetImportException(
            _(u"Unknown data type '%s' ") % (type))
    try:
        return converters[type](value)
    except:
        raise SheetImportException(
            _(u"Can not process value '%s' of type '%s' ") %
            (value, type))


def normalize_row(raw_row):
    # Unify difference between CSV and openpyxl cells
    return [getattr(v, 'value', v) for v in raw_row]


def process_rows(rows, profile_columns, skip_header=False):
    # If there is no header (skip_header=False), then use profile's order of
    # columns, otherwise use header line to check mapping and define order
    first_row = normalize_row(rows.next()) if skip_header else None
    columns = order_columns(profile_columns, first_row)
    # columns = [{'field': "...", 'type': "..."}, ...]

    objects = []
    errors = []
    for i, row in enumerate(rows, 2 if first_row else 1):
        try:
            objects.append(process_row(row, columns))
        except SheetImportException as e:
            raise type(e), type(e)(e.message +
                                   'in row %d ' % i), sys.exc_info()[2]

    return objects


def process_row(row, columns):
    values = normalize_row(row)
    return reduce(
        lambda object_dict, converter: converter.add_to(object_dict),
        map(CellConverter, values, columns),
        {}
    )


class CellConverter(object):
    def __init__(self, value, col_spec):
        self.value = value
        self.type = col_spec['type']
        self.field = col_spec['field']

    def add_to(self, object_dict):
        if self.type != 'ignore':
            object_dict[self.field] = convert_value(
                self.value, self.type)
        return object_dict


def save_rows(objects, data_type):
    for obj in objects:
        create_message(obj)

    return len(objects)


def store_spreadsheet(label, fobject):
    profile = get_profile(label)

    file_format = profile.get('format')
    skip_header = profile.get('skip_header', False)

    rows = get_rows_iterator(fobject, file_format)
    items = process_rows(rows, profile['columns'], skip_header)
    return save_rows(items, 'message')
