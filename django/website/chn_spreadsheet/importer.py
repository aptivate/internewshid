import dateutil.parser
from decimal import Decimal
import datetime
import pytz
import sys

from django.utils.timezone import is_naive
from django.utils.translation import ugettext as _
from openpyxl import load_workbook

import transport

from .models import SheetProfile


class SheetImportException(Exception):
    pass


class Importer(object):

    def get_profile(self, label):
        try:
            sheet_profile = SheetProfile.objects.get(label=label)
        except SheetProfile.DoesNotExist:
            error_msg = _('Misconfigured service. Source "%s" does not exist') % label
            raise SheetImportException(error_msg)

        return sheet_profile.profile

    def get_columns_map(self, col_list):
        '''This function assumes that column names are unique for spreadsheet.
        If they are not, then you already have a problem.'''

        return {column['name']: column for column in col_list}

    def get_rows_iterator(self, spreadsheet, file_format):
        if file_format == 'excel':
            try:
                wb = load_workbook(spreadsheet, read_only=True)
                ws = wb[wb.sheetnames[0]]
            except:
                error_msg = _('Expected excel file. Received file in an unrecognized format.')
                raise SheetImportException(error_msg)
            rows = ws.rows
        else:
            error_msg = _('Unsupported file format: %s') % file_format
            raise SheetImportException(error_msg)

        return rows

    def order_columns(self, profile_columns, first_row=None):
        columns = []
        if first_row:
            col_map = self.get_columns_map(profile_columns)

            for label in first_row[:len(col_map)]:
                try:
                    columns.append(col_map[label])
                except:
                    error_msg = _('Unknown column: %s') % label
                    raise SheetImportException(error_msg)
        else:
            columns = [d.copy() for d in profile_columns]

        return columns

    def get_fields_and_types(self, columns):
        fields = [col['field'] for col in columns]
        types = [col['type'] for col in columns]

        return fields, types

    def normalize_row(self, raw_row):
        # Unify difference between CSV and openpyxl cells
        return [getattr(v, 'value', v) for v in raw_row]

    def process_rows(self, rows, profile_columns, meta_data, skip_header=False):
        # If there is no header (skip_header=False), then use profile's order of
        # columns, otherwise use header line to check mapping and define order
        first_row = self.normalize_row(rows.next()) if skip_header else None
        columns = self.order_columns(profile_columns, first_row)
        # columns = [{'field': "...", 'type': "..."}, ...]

        objects = []
        for i, row in enumerate(rows, 2 if first_row else 1):
            try:
                values = self.normalize_row(row)

                if any(values):
                    item = self.process_row(values, columns)

                    for taxonomy, term in meta_data.iteritems():
                        self._append_term_to_item(item, taxonomy, term)

                    objects.append(item)

            except SheetImportException as e:
                raise type(e), type(e)(e.message +
                                       'in row %d ' % i), sys.exc_info()[2]

        return objects

    def process_row(self, values, columns):
        item = {}

        for val, col in zip(values, columns):
            converter = CellConverter(val, col)

            if col['type'] == 'taxonomy':
                self._append_term_to_item(
                    item, col['taxonomy'], converter.convert_value())
            else:
                converter.add_to(item)

        return item

    def _append_term_to_item(self, item, taxonomy, name):
        term = self._get_term_dict(taxonomy, name)
        item.setdefault('terms', []).append(term)

    def _get_term_dict(self, taxonomy, name):
        return {'taxonomy': taxonomy, 'name': name}

    def save_rows(self, objects):
        for obj in objects:
            terms = obj.pop('terms')
            item = transport.items.create(obj)
            for term in terms:
                transport.items.add_terms(
                    item['id'], term['taxonomy'], term['name'])

        return len(objects)

    def store_spreadsheet(self, label, fobject):
        profile = self.get_profile(label)

        file_format = profile.get('format')
        skip_header = profile.get('skip_header', False)
        meta_data = profile.get('taxonomies')

        rows = self.get_rows_iterator(fobject, file_format)

        items = self.process_rows(rows, profile['columns'], meta_data,
                                  skip_header)

        return self.save_rows(items)


class CellConverter(object):
    def __init__(self, value, col_spec):
        self.value = value
        self.type = col_spec['type']
        self.field = col_spec['field']
        self.date_format = col_spec.get('date_format', None)

    def add_to(self, object_dict):
        if self.type != 'ignore':
            object_dict[self.field] = self.convert_value()
        return object_dict

    def convert_value(self):
        converters = {
            'date': lambda x: self.convert_date(),
            'text': lambda x: x,
            'integer': lambda x: int(x),
            'number': lambda x: Decimal(x),
            'taxonomy': lambda x: x,
        }
        if self.type not in converters:
            raise SheetImportException(
                _(u"Unknown data type '%s' ") % (self.type))
        try:
            return converters[self.type](self.value)
        except Exception as e:
            message = _("%s\nCan not process value '%s' of type '%s' ") % (e.message, self.value, self.type)
            raise SheetImportException(message), None, sys.exc_info()[2]

    def convert_date(self):
        if self.value is None:
            return None

        if isinstance(self.value, basestring):
            date_time = self.parse_date()
        else:
            date_time = self.value

        if is_naive(date_time):
            date_time = pytz.utc.localize(date_time)

        return date_time

    def parse_date(self):
        if self.date_format is None:
            raise SheetImportException(
                _(u"Date format not specified for '%s' ") %
                (self.field))

        try:
            date_time = datetime.datetime.strptime(self.value,
                                                   self.date_format)
        except:
            date_time = dateutil.parser.parse(self.value)

        return date_time
